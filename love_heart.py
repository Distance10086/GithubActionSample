import socket
import threading
import tkinter as tk
from tkinter import ttk, filedialog
from datetime import datetime
import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.styles import Alignment
import re
import os
import time

DEFAULT_DNS = "223.5.5.5"
DEFAULT_DOMAIN_FILE = os.path.join(os.getcwd(), "domain.txt")
LOG_START = "DNS解析开始：{}\n"
LOG_END = "DNS解析结束：{}\n"
LOG_PROCESSING_START = "Excel处理开始...\n"
LOG_PROCESSING_END = "Excel处理完成，文件保存在{}\n"
ERROR_MESSAGE = "解析失败"

def get_ip_addresses(domain, dns_server):
    try:
        resolver = socket.gethostbyname_ex(domain)
        ip_addresses = resolver[2]
        return "\n".join(ip_addresses)
    except socket.gaierror:
        return ERROR_MESSAGE

def read_domains(file_path):
    try:
        with open(file_path, 'r') as file:
            return [line.strip() for line in file]
    except FileNotFoundError:
        return None

def clean_ip_addresses(ip_addresses):
    return re.sub(r"[\"']", "", ip_addresses)

def extract_c_segment(ip_addresses):
    c_segments = []
    for ip in ip_addresses.split('\n'):
        parts = ip.split('.')
        if len(parts) >= 3:
            c_segment = '.'.join(parts[:3])
            c_segments.append(c_segment)
    return "\n".join(c_segments)

def adjust_column_width(sheet):
    for column in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value)
        adjusted_width = max_length + 2
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

def write_to_excel(results, output_path):
    workbook = openpyxl.Workbook()
    sheet1 = workbook.active
    sheet1.title = "Domain IPs"
    sheet1.append(["域名", "IP地址", "C段"])

    for domain, ip_addresses in results:
        ip_addresses = clean_ip_addresses(ip_addresses)
        c_segment = clean_ip_addresses(extract_c_segment(ip_addresses))
        sheet1.append([domain, ip_addresses, c_segment])
        current_row = sheet1.max_row
        sheet1.cell(row=current_row, column=2).alignment = Alignment(wrap_text=True)
        sheet1.cell(row=current_row, column=3).alignment = Alignment(wrap_text=True)

    adjust_column_width(sheet1)

    sheet2 = workbook.create_sheet(title="Domain IPs Detail")
    sheet2.append(["域名", "IP地址", "C段"])

    for domain, ip_addresses in results:
        ip_addresses = clean_ip_addresses(ip_addresses)
        for ip in ip_addresses.split('\n'):
            c_segment = extract_c_segment(ip)
            sheet2.append([domain, ip, c_segment])

    adjust_column_width(sheet2)
    workbook.save(output_path)

def resolve_domains(domains, dns_server, progress_callback, log_callback, stop_event):
    results = []
    total = len(domains)
    start_time = time.time()

    with ThreadPoolExecutor(max_workers=200) as executor:
        future_to_domain = {executor.submit(get_ip_addresses, domain, dns_server): domain for domain in domains}
        for i, future in enumerate(as_completed(future_to_domain)):
            if stop_event.is_set():
                break
            domain = future_to_domain[future]
            try:
                ip_addresses = future.result()
                results.append((domain, ip_addresses))
            except Exception as e:
                results.append((domain, f"{ERROR_MESSAGE}: {e}"))
            progress_callback(i + 1, total)
            log_callback(i + 1, total, time.time() - start_time)

    elapsed_time = time.time() - start_time
    log_callback(total, total, elapsed_time)
    return results

class DomainResolverApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("域名批量解析")
        self.geometry("600x350")
        self.dns_server = tk.StringVar(value=DEFAULT_DNS)
        self.domain_file_path = tk.StringVar(value=DEFAULT_DOMAIN_FILE)
        self.stop_event = threading.Event()
        self.create_widgets()

    def create_widgets(self):
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(1, weight=1)

        tk.Label(self, text="DNS 服务器:", anchor='w').grid(row=0, column=0, sticky='w')
        tk.Entry(self, textvariable=self.dns_server, width=30).grid(row=0, column=1, sticky='ew')

        domain_frame = tk.Frame(self)
        domain_frame.grid(row=1, column=0, columnspan=2, sticky='ew', padx=10, pady=5)
        domain_frame.grid_columnconfigure(1, weight=1)

        tk.Label(domain_frame, text="域名文件:   ", anchor='w').grid(row=0, column=0, sticky='w')
        tk.Entry(domain_frame, textvariable=self.domain_file_path, width=30).grid(row=0, column=1, sticky='ew')
        tk.Button(domain_frame, text="选择域名文件", command=self.choose_domain_file).grid(row=0, column=2, sticky='ew', padx=(5, 0))

        progress_frame = tk.Frame(self)
        progress_frame.grid(row=2, column=0, columnspan=2, sticky='ew', padx=10, pady=(0, 10))
        self.progress_label_left = tk.Label(progress_frame, text="0%")
        self.progress_label_left.pack(side=tk.LEFT)
        self.progress = ttk.Progressbar(progress_frame, mode="determinate")
        self.progress.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 5))
        self.progress_label_right = tk.Label(progress_frame, text="0/0")
        self.progress_label_right.pack(side=tk.RIGHT)

        button_frame = tk.Frame(self)
        button_frame.grid(row=3, column=0, columnspan=2, sticky='ew', pady=10)
        button_frame.grid_columnconfigure(0, weight=1)
        button_frame.grid_columnconfigure(1, weight=1)
        tk.Button(button_frame, text="开始解析", command=self.start_resolving).grid(row=0, column=0, padx=5)
        tk.Button(button_frame, text="停止解析", command=self.stop_resolving).grid(row=0, column=1, padx=5)

        self.log_text = tk.Text(self, height=10)
        self.log_text.grid(row=4, column=0, columnspan=2, sticky='nsew', padx=10, pady=(0, 5))

    def log_message(self, completed, total, elapsed_time, end=False):
        self.log_text.tag_configure("bold", font=("TkDefaultFont", 10, "bold"))
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if completed == 0:
            self.log_text.insert(tk.END, LOG_START.format(current_time), "bold")
        elif end:
            self.log_text.insert(tk.END, LOG_END.format(current_time), "bold")
            elapsed_msg = f"DNS解析耗时: {elapsed_time:.2f}秒\n"
            self.log_text.insert(tk.END, elapsed_msg, "bold")
        self.log_text.see(tk.END)

    def choose_domain_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Text Files", "*.txt")])
        if file_path:
            self.domain_file_path.set(file_path)

    def start_resolving(self):
        domain_file = self.domain_file_path.get()
        if not domain_file:
            self.log_text.insert(tk.END, "错误: 请选择域名文件\n", "bold")
            return

        domains = read_domains(domain_file)
        if domains is None:
            self.log_text.insert(tk.END, f"文件不存在: {domain_file}\n", "bold")
            return

        dns_server = self.dns_server.get()
        self.stop_event.clear()

        current_time = datetime.now().strftime("%Y%m%d%H%M%S")
        self.output_file = f'domain_ips_{current_time}.xlsx'

        total = len(domains)
        self.progress["maximum"] = total
        self.progress["value"] = 0
        self.progress_label_left["text"] = "0%"
        self.progress_label_right["text"] = f"0/{total}"

        def progress_callback(completed, total):
            self.progress["value"] = completed
            percentage = (completed / total) * 100
            self.progress_label_left["text"] = f"{percentage:.0f}%"
            self.progress_label_right["text"] = f"{completed}/{total}"

        def run():
            start_time = time.time()
            self.log_message(0, total, 0)
            results = resolve_domains(domains, dns_server, progress_callback, self.log_message, self.stop_event)
            if not self.stop_event.is_set():
                self.log_message(total, total, time.time() - start_time, end=True)
                self.log_text.insert(tk.END, LOG_PROCESSING_START, "bold")
                write_to_excel(results, self.output_file)
                self.log_text.insert(tk.END, LOG_PROCESSING_END.format(self.output_file), "bold")

        threading.Thread(target=run).start()

    def stop_resolving(self):
        self.stop_event.set()
        self.progress["value"] = 0
        self.progress_label_left["text"] = "0%"
        self.progress_label_right["text"] = "0/0"
        self.log_text.insert(tk.END, "解析已停止\n", "bold")

if __name__ == "__main__":
    app = DomainResolverApp()
    app.mainloop()
